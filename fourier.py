import numpy as np
from scipy.fft import fft, fftfreq
import scipy.io.wavfile as wavfile
import openpyxl
import math
import random
import os

os.chdir('..')

BINSIZE = 100
MAXSAMPLESAMOUNT = 25000000
MAXFREQUENCY = 25000
    
def fourierTransformParse(sampleRate, tone, samplesAmount):
    xlsxFilePath = "fourier.xlsx"
    workbook = openpyxl.load_workbook(xlsxFilePath)
    
    sheet = workbook.active
    sheet.delete_cols(1,2)
    
    print('Dataset size:', samplesAmount, 'items')
    print(round(samplesAmount / MAXSAMPLESAMOUNT), 'chunks of size:', MAXSAMPLESAMOUNT)
    
    bins = dict(list(enumerate([0]*round(MAXFREQUENCY / BINSIZE + 1))))
    
    print(len(bins), 'bins of width:', BINSIZE)
    print('Binning data...')
    
    for i in range(0, samplesAmount, MAXSAMPLESAMOUNT):
        print('Chunk', round(i/MAXSAMPLESAMOUNT) + 1, 'of', round(samplesAmount / MAXSAMPLESAMOUNT))
        toneChunk = tone[i: MAXSAMPLESAMOUNT+i]
        
        chunkFrequencies = fftfreq(MAXSAMPLESAMOUNT, 1 / sampleRate)
        occurances = fft(toneChunk)
        
        bins = writeDataToBin(chunkFrequencies, occurances, sheet, bins)
    
    print('Writing to data sheet...')
    for bin in bins.items():
        sheet.cell(row = bin[0]+1, column = 1).value = bin[0]
        sheet.cell(row = bin[0]+1, column = 2).value = (math.log(bin[1], 10) if bin[1] > 0 else 0)
    
    print('Saving to xlsx file...')
    workbook.save(xlsxFilePath)

def writeDataToBin(chunkFrequencies, occurances, sheet, bins):
    chunkSize = len(occurances)

    for i, frequency in enumerate(chunkFrequencies):
        if frequency > MAXFREQUENCY or i >= len(occurances): continue
        
        binNumber = int(abs(frequency / BINSIZE))
        bins[binNumber] += int(np.abs(occurances[i].real[0]))
        
        if random.choice(range(chunkSize)) < 10: print(round(i / chunkSize * 100), 'percent')
        
    return bins

print('reading file... Do not ^C now')
sampleRate, tone = wavfile.read('B304.wav')
samplesAmount = tone.shape[0]

fourierTransformParse(sampleRate, tone, samplesAmount)
print('Done!')